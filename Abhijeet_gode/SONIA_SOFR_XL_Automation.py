import os
import time
import random
import calendar
import openpyxl
import datetime
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
import win32com.client as win32
from openpyxl import load_workbook
from datetime import date, timedelta
from selenium.webdriver.common.by import By
from openpyxl.utils import get_column_letter
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import ElementNotInteractableException
fipath = input("\nEnter the file path \U0001F917: ")
file = pd.read_excel(r'{0}'.format(fipath))
print("\nAutomation Under progress, Please wait...\U0001FAE1")
options = webdriver.ChromeOptions()
prefs = {"download.default_directory" : "C:\Downloads"}
options.add_experimental_option("prefs", prefs)
options.add_argument("--start-maximized")
service = Service(executable_path='C:\Chromedriver\chromedriver')
driver = webdriver.Chrome(service=service, options=options)
date= file['Unnamed: 0'].tail(1).to_string()
date = date[8:18]
da = int(date[5:7])
da = datetime.date(1900, da, 1).strftime('%b')
print("\n\nWebsite data captures started for SOFR... \U0001F929\n")
SOFR = driver.get('https://fred.stlouisfed.org/series/SOFR')
fd = date
time.sleep(3)
frmd = driver.find_element(By.XPATH, '//*[@id="input-cosd"]').clear()
time.sleep(1)
frmd = driver.find_element(By.XPATH, '//*[@id="input-cosd"]')
inputEle = frmd.send_keys(fd)
downloadexl = driver.find_element(By.XPATH, '//*[@id="download-button"]')
downloadexl.click()
driver.implicitly_wait(2)
excelopt = driver.find_element(By.XPATH, '//*[@id="download-data"]')
excelopt.click()
time.sleep(2)
print("\n\nWebsite data captures started for SONIA... \U0001F60D\n")
Sonia = driver.get('https://www.bankofengland.co.uk/boeapps/database/fromshowcolumns.asp?Travel=NIx&ShadowPage=1&SearchText=sonia&SearchExclude=&SearchTextFields=TC&Thes=&SearchType=&Cats=&ActualResNumPerPage=&TotalNumResults=12&XNotes2=Y&C=5JK&ShowData.x=41&ShowData.y=34')#("window.open")
time.sleep(2)
search_btn = driver.find_element(By.XPATH, '//html//body//div[2]//div[3]//div//div//table//tbody//tr[2]//td[3]//button')
try:
    search_btn.click()
except:
    ElementNotInteractableException
time.sleep(2)
from_d1 = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/section[1]/div/form/div[2]/div[1]/table/tbody/tr[1]/td[2]/select[1]").click()
from_d1 = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/section[1]/div/form/div[2]/div[1]/table/tbody/tr[1]/td[2]/select[1]")
inp = from_d1.send_keys(date[8:10])
time.sleep(0.5)
from_d2 = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/section[1]/div/form/div[2]/div[1]/table/tbody/tr[1]/td[2]/select[2]").click()
from_d2 = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/section[1]/div/form/div[2]/div[1]/table/tbody/tr[1]/td[2]/select[2]")
inp = from_d2.send_keys(da)
time.sleep(0.5)
from_d3 = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/section[1]/div/form/div[2]/div[1]/table/tbody/tr[1]/td[2]/select[3]").click()
from_d3 = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/section[1]/div/form/div[2]/div[1]/table/tbody/tr[1]/td[2]/select[3]")
inp = from_d3.send_keys(date[:4])
time.sleep(0.5) 
viewD = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/section[1]/div/form/div[3]/input").click()
time.sleep(2)
excel = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/section[2]/div/div[1]/div[1]/a[3]").click()
time.sleep(1)
print("\nWebsite data captures started for Term SOFR_Rates is saving... \U0001F973\n")
Term_SOFR_Rates = driver.get("https://www.cmegroup.com/market-data/cme-group-benchmark-administration/term-sofr.html")
time.sleep(2)
search_btn = driver.find_element(By.XPATH, '/html/body/div[4]/div[2]/div/div[1]/div/div[2]/div/button[2]')
try:
    search_btn.click()
except:
    ElementNotInteractableException
html = driver.page_source
table = pd.read_html(html)[0]
number = random.randint(0, 100)
table.to_excel(f'C:\\Downloads\\Term_SOFR_Rates_{number}.xlsx')
print("\nPlease Work around with ICE website by click on Google Chrome browser... \U0001F927\n")
ne = driver.get("https://www.theice.com/marketdata/reports/276")
##############################################################################################################################################
print("\nFile Operation is started now please wait for the confirmation.... \U0001F609 \U0001F607\n")
fname = "C:\Downloads\SOFR.xls"
excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Visible=False
wkb = excel.Workbooks.Open(fname)
wkb.SaveAs(fname+"x", FileFormat = 51)
wkb.Close()
excel.Application.Quit()
file_s = r'C:\Downloads\SOFR.xlsx'
df = pd.read_excel(file_s)
df.fillna(0)
df.to_excel(file_s, index=False)
file_name = "C:\Downloads\Bank of England  Database.xlsx"
excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Visible=False
wkb = excel.Workbooks.Open(file_name)
wkb.SaveAs(file_name, FileFormat = 51)
wkb.Close()
excel.Application.Quit()
d = pd.read_excel(file_name)
a = d.drop(index=0)[::-1]
a.to_excel(file_name, header=None, index=False)
file = fipath
wb = load_workbook(file)
ws = wb.worksheets[0]
wss = wb.worksheets[1]
file_s = r'C:\Downloads\SOFR.xlsx'
wb1 = load_workbook(file_s)
ws1 = wb1.worksheets[0]
file1 = r'C:\Downloads\Bank of England  Database.xlsx'
wb2 = openpyxl.load_workbook(file1)
ws2 = wb2.worksheets[0]
ws1.delete_rows(1,11)
for col in range(1, 10):
    max_col_row1 = len([cell for cell in ws["A"] if cell.value])
for col in range(1, 10):
    max_col_row2 = len([cell for cell in wss['A'] if cell.value])
max_r1 = ws1.max_row
max_c1 = ws1.max_column
max_r2 = ws2.max_row
max_c2 = ws2.max_column
for i in  range(1, max_r1+1):
    for j in  range(1, max_c1+1):
        c = ws1.cell(row=i, column=j)
        d = ws.cell(row=i+max_col_row1+1, column=j).value = c.value
for i in range(1, max_r2+1)[::-1]:
    for j in range(1, max_c2+1):
        a = ws2.cell(row=i, column=j)
        b = wss.cell(row=i+max_col_row2+1, column=j).value = a.value
for i in range(1, max_r1+1):
        n = '=B{0}/100'.format(i+max_col_row1)
        m = '=1+C{0}/250'.format(i+max_col_row1)
        o = '=E{0}+1'.format(i+max_col_row1)
        p = '=(E{0}=1)*D{0}+(E{0}>1)*F{1}*D{0}'.format(i+max_col_row1+1, i+max_col_row1)
        q = "=--(E{0}>=MIN(E:E)+20)".format(i+max_col_row1+1)
        r = "=--(E{0}>=MIN(E:E)+62)".format(i+max_col_row1+1)
        s = "=--(E{0}>=MIN(E:E)+125)".format(i+max_col_row1+1)
        t = "=--(E{0}>=MIN(E:E)+251)".format(i+max_col_row1+1)
        u = "=((F{0}/F{1}-1)*12)/G{0}".format(i+max_col_row1+1, max_col_row1-20+i)
        v = "=((F{0}/F{1}-1)*4)/H{0}".format(i+max_col_row1+1, max_col_row1-62+i)
        w = "=((F{0}/F{1}-1)*2)/I{0}".format(i+max_col_row1+1, max_col_row1-125+i)
        x = "=((F{0}/F{1}-1)*1)/J{0}".format(i+max_col_row1+1, max_col_row1-251+i)
        ws.cell(row=i+max_col_row1+1, column=3).value = n
        ws.cell(row=i+max_col_row1+1, column=4).value = m
        ws.cell(row=i+max_col_row1+1, column=5).value = o
        ws.cell(row=i+max_col_row1+1, column=6).value = p
        ws.cell(row=i+max_col_row1+1, column=7).value = q
        ws.cell(row=i+max_col_row1+1, column=8).value = r
        ws.cell(row=i+max_col_row1+1, column=9).value = s
        ws.cell(row=i+max_col_row1+1, column=10).value = t
        ws.cell(row=i+max_col_row1+1, column=11).value = u
        ws.cell(row=i+max_col_row1+1, column=12).value = v
        ws.cell(row=i+max_col_row1+1, column=13).value = w
        ws.cell(row=i+max_col_row1+1, column=14).value = x
for i in range(1, max_r2+1):
        n = '=B{0}/100'.format(i+max_col_row2)
        m = '=1+C{0}/250'.format(i+max_col_row2)
        o = '=E{0}+1'.format(i+max_col_row2)
        p = '=(E{0}=1)*D{0}+(E{0}>1)*F{1}*D{0}'.format(i+max_col_row2+1, i+max_col_row2)
        q = "=--(E{0}>=MIN(E:E)+20)".format(i+max_col_row2+1)
        r = "=--(E{0}>=MIN(E:E)+62)".format(i+max_col_row2+1)
        s = "=--(E{0}>=MIN(E:E)+125)".format(i+max_col_row2+1)
        t = "=--(E{0}>=MIN(E:E)+251)".format(i+max_col_row2+1)
        u = "=((F{0}/F{1}-1)*12)/G{0}".format(i+max_col_row2+1, max_col_row2-20+i)
        v = "=((F{0}/F{1}-1)*4)/H{0}".format(i+max_col_row2+1, max_col_row2-62+i)
        w = "=((F{0}/F{1}-1)*2)/I{0}".format(i+max_col_row2+1, max_col_row2-125+i)
        x = "=((F{0}/F{1}-1)*1)/J{0}".format(i+max_col_row2+1, max_col_row2-251+i)
        wss.cell(row=i+max_col_row2+1, column=3).value = n
        wss.cell(row=i+max_col_row2+1, column=4).value = m
        wss.cell(row=i+max_col_row2+1, column=5).value = o
        wss.cell(row=i+max_col_row2+1, column=6).value = p
        wss.cell(row=i+max_col_row2+1, column=7).value = q
        wss.cell(row=i+max_col_row2+1, column=8).value = r
        wss.cell(row=i+max_col_row2+1, column=9).value = s
        wss.cell(row=i+max_col_row2+1, column=10).value = t
        wss.cell(row=i+max_col_row2+1, column=11).value = u
        wss.cell(row=i+max_col_row2+1, column=12).value = v
        wss.cell(row=i+max_col_row2+1, column=13).value = w
        wss.cell(row=i+max_col_row2+1, column=14).value = x
for i in range(1, max_r2+1):
        for c in range(11, 15):
            wss.cell(row=i+max_col_row2+1, column=c).number_format= '0.00%'
            wss.cell(row=i+max_col_row2+1, column=3).number_format= '0.00%'
print("\n\nFile Operation is Completed Now wait for the file to save.... \U0001F4A5 \U0001F4A5 \U0001F4A5\n")
os.remove(fname)
os.remove(file_s)
os.remove(file_name)
wb.save('C:\Downloads\Output_of_{}.xlsx'.format(fipath[:-5]))
print("\nAutomation is Successful You can close the program... \U0001F9DE \U0001F60B\n")
print("Thank You... ")