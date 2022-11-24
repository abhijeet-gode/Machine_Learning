#Importing important libraries
print("Application is started please wait for input.................")
import pandas as pd
pd.set_option("display.max_colwidth", None)
import numpy as np
import getpass
import os
import datetime
from datetime import date
from datetime import datetime
import re
import glob
import openpyxl
import cx_Oracle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Color, colors, PatternFill, Font, Border
from openpyxl.cell import cell
from openpyxl.styles import Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension
try:
        from openpyxl.cell import get_column_letter
except ImportError:
        from openpyxl.utils import get_column_letter
#********************************************************************************************************
user = getpass.getuser()
desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Box')
drive = desktop[0]
# file_name = input("Enter the file : ")
# BAU_p = r'C:\Users\{0}\Box\EMEA Mumbai Data Operations\BAU 2022\{1}.xlsx'.format(user,file_name)
BAU_p = r'C:\Users\abhijeet_gode\Downloads\Abhijeet_Gode1\Bond Map\BAU_sep_2022.xlsx'
BAU_Sheet = pd.read_excel(BAU_p,sheet_name='Main',usecols=['Deal Name','Core Id','Cut-off date','Asset Class','Ticker Series','Deal Status','Processor', 'Deal Type'])
Dash = pd.read_excel("C:\\Users\\"+ str(getpass.getuser()) +"\\Box\\EMEA FDM Dashbroad\\SF_EMEA_Dashboard.xlsm", sheet_name = "Consolidate", header = 2)
Dashboard = Dash.reset_index(drop=True)
Proces = input("Enter the name of Processor : ").title()
Deal_status = input("Enter the Deal Status : ").title()
Deal_type = input("Enter the Deal Type : ").title()
#BAU Filter
Bau_ = BAU_Sheet.loc[(BAU_Sheet['Processor'] == Proces) & (BAU_Sheet['Deal Status'] == Deal_status) & (BAU_Sheet['Deal Type'] == Deal_type)]
Bau_filter = Bau_.reset_index(drop=True)
print(Bau_filter)
core_ = Bau_filter['Core Id']
c = int(input("Enter the index of value 0 to 1 :"))
core_id = core_[c]
coff_date = Bau_filter.loc[Bau_filter['Core Id'] == core_id,'Cut-off date']
deal_name = Bau_filter.loc[Bau_filter['Core Id'] == core_id, 'Deal Name'].to_string()
asset_class = Bau_filter.loc[Bau_filter['Core Id'] == core_id, 'Asset Class'].to_string()
ticker = Bau_filter.loc[Bau_filter['Core Id'] == core_id, 'Ticker Series'].tolist()
split = (re.findall(r"[^\W\d_]+|\d+",ticker[c]))
full = ticker
full
#SQL
dsn_tns = cx_Oracle.makedsn(r'fgraci31aprod3-scan.pde.spratingsvpc.com', '1521', service_name=r'SFREFPRD.WORLD')
cnn = cx_Oracle.connect(user='abhijeet_gode', password='Abhi#0311', dsn=dsn_tns)
inpc = int(input("Enter the core ID : "))
string_sql = r'''Select 
b.Deal_ID as core_id
,a.AS_OF_DATE as asofdt
,a.CUR_BAL_AMT as cbal
,a.CUR_COLLAT_CNT as count
, a.pay_month - TO_CHAR(a.AS_OF_DATE,'YYYYMM') as OFFSET

from sfmstr.VW_DEAL_PERF_EMEA_CRISIL a, SFMSTR.VW_DEAL_ID_LINK b

Where 
a.SP_DEAL_CODE = b.Parent_sp_Deal_code
and b.deal_id_type = 'COREISS'
and b.deal_id = {}
order by a.AS_OF_DATE asc'''.format(inpc)
srl_sql = (string_sql)
SRL = pd.read_sql(srl_sql,cnn)
sqld = SRL["ASOFDT"].max()
sqldt = str(sqld)[:10]

CBAL = SRL.loc[SRL['ASOFDT'] == sqldt,'CBAL']
COUNT = SRL.loc[SRL['ASOFDT'] == sqldt,'COUNT']
OFFSET = SRL.loc[SRL['ASOFDT'] == sqldt,'OFFSET']
#Aggregate Data Map
#File imported for processing
Aggre_map = "C:\\Users\\{0}\\Box\\Aggregate Data Maps\\".format(user)
Aggregate_map = glob.glob(r'C:\\Users\\{0}\\Box\\Aggregate Data Maps\\{1}*\\{2}*\\*.map'.format(user, split[0], full[:5]), recursive=False)
for i in Aggregate_map:
    mapp = pd.read_table(i)

# XML asofdt mapping	
first_agg = mapp[(mapp['SECTION NAME'] == 'HEADER_REC') & (mapp['FIELD NAME'] == 'ASOFDT')]
asofdt_mapping = first_agg["SOURCE"].to_string()
# Aggregate CBAL mapping
second_agg = mapp[(mapp['SECTION NAME'] == 'COLLATERAL_REC') & (mapp['FIELD NAME'] == 'CBAL')]
aggre_CBAL = second_agg["SOURCE"].to_string()
# Aggregate Count mapping
third_agg = mapp[(mapp['SECTION NAME'] == 'COLLATERAL_REC') & (mapp['FIELD NAME'] == 'COUNT')]
aggre_count = third_agg["SOURCE"].to_string()

#Extra Value
date = Bau_filter[['Cut-off date']]
date1 = date["Cut-off date"].to_string(index=False).split('\n')
datem = datetime.strptime(date1[c], "%Y-%m-%d")
dateo=datem.month
if dateo in [1, 4, 7, 10]:
    qua ='Jan-Apr-Jul-Oct'
elif dateo in [2, 5, 8, 11]:
    qua='Feb-May-Aug-Nov'
else:# dateo in [3, 6, 9, 12]:
    qua = 'Mar-Jun-Sep-Dec'

coff_date1 = coff_date.to_string(index=False).split('\n')
date_1 = coff_date1[0]
date_2 = sqldt
datem = datetime.strptime(date_1, "%Y-%m-%d")
daten = datetime.strptime(date_2, "%Y-%m-%d")
dateof = datem.month
datein = daten.month
if dateof == datein:
    euq = 'Cuttoffdt equal asofdt'
else:
    euq = 'Cuttoffdt plus 1 asofdt'

if qua in 'Jan-Apr-Jul-Oct':
    neww = [1, 4, 7, 10]
elif qua in 'Feb-May-Aug-Nov':
    neww= [2, 5, 8, 11]
else:# dateo in [3, 6, 9, 12]:
    neww = [3, 6, 9, 12]

if dateof in neww:
    a = 0
else:
    a = 1

#Dashboard Sorting
Dash_frequency = Dashboard[Dashboard['Core Issue ID ']==inpc]
LL_frequency = Dash_frequency['LL Reporting frequency'].to_string()
Agg_frequency = Dash_frequency['Agg Reporting frequency'].to_string()
#Loan
dt = sqldt.replace('-', '')
path = glob.glob(r"W:\\RMBS\\{0}\\Performance\\{1}*\\Clean\\{2}*.csv".format(inpc, dt, deal_name[5:]), recursive=False)
for i in path:
    RMBS = pd.read_csv(i, usecols=['CURR_DELINQ_STAT_CD', 'CBAL_AMT'])
first_w = RMBS[(RMBS['CURR_DELINQ_STAT_CD'] == 'RED') & (RMBS['CBAL_AMT'])]
total_cbal = sum(first_w['CBAL_AMT'])
cbal_count = first_w['CBAL_AMT'].count()
diff_cou = COUNT-cbal_count
differ = CBAL-total_cbal
diff_count = diff_cou.to_string()[5:]
diff_cbal = differ.to_string()[5:]
maxcbal = float(CBAL) if float(CBAL) > total_cbal else total_cbal
maxcount = int(COUNT) if int(COUNT) > cbal_count else cbal_count
different_count = int(diff_cou)/maxcount*100
different_cbal = float(differ)/maxcbal*100

your_workbook = Workbook()    #creating the workbook
sheet = your_workbook.active
sheet.merge_cells('A1:B1')
sheet["A1"] = "FORMAT"
sheet["A2"] = "Deal Name (Full)"
sheet["B2"] = str(deal_name[5:])
sheet["A3"] = "Pool cutoff date"
sheet["B3"] = str(coff_date1[0])
sheet["A4"] = "XML asofdt"
sheet["B4"] = str(sqldt)
sheet["A5"] = "XML asofdt mapping"
sheet["B5"] = str(asofdt_mapping[5:])
sheet["A6"] = "XML offset status"
sheet["B6"] = str(a)
sheet["A7"] = "Asofdt mapping suggestion"
sheet["B7"] = str(euq)
sheet["A8"] = "Aggregate CBAL mapping"
sheet["B8"] = str(aggre_CBAL[6:])
sheet["A9"] = "Aggregate Count mapping"
sheet["B9"] = str(aggre_count[6:])
sheet["A10"] = "LL frequency"
sheet["B10"] = str(LL_frequency[7:])
sheet["A11"] = "Aggregate Frequency"
sheet["B11"] = str(Agg_frequency[7:])
sheet["A12"] = "Bond Payment frequency"
sheet["B12"] = str(qua)
sheet["A13"] = "Pay Month(Y/N)"
sheet["B13"] = "Y"
if Deal_type=="Third Performance":
    sheet["A15"] = "Period – Sep 2021"
    sheet["B15"] = "Third Performance"
    sheet["A17"] = ""
    sheet["B17"] = "XML"
    sheet["B18"] = float(CBAL)
    sheet["B19"] = int(COUNT)
    sheet["C18"] = str(total_cbal)
    sheet["C19"] = str(cbal_count)
    sheet["D18"] = str(diff_cbal)
    sheet["D19"] = str(diff_count)
    sheet["E18"] = str(different_cbal)
    sheet["E19"] = str(different_count)
    sheet["C17"] = "Loan Level"
    sheet["D17"] = "Difference"
    sheet["E17"] = "Difference %"
    sheet["A18"] = "CBAL"
    sheet["A19"] = "COUNT*"
    regular = Side(border_style="medium", color="000000")
    no_left_side = Border(top = regular,bottom=regular,right=regular)
    for c in sheet['A17:A19']+sheet['B17:B19']+sheet['C17:C19']+sheet['D17:D19']+sheet['E17:E19']:
        c[0].border = no_left_side
    for b in sheet['A17:A19']+sheet['B15:B17']+sheet['C17:C17']+sheet['D17:D17']+sheet['E17:E17']:
        b[0].font = Font(bold=True)
else:
    pass
if Deal_type=="Second Performance":
    sheet["A15"] = "Period – Sep 2021"
    sheet["B15"] = "Second Performance"
    sheet["A17"] = ""
    sheet["B17"] = "XML"
    sheet["B18"] = float(CBAL)
    sheet["B19"] = int(COUNT)
    sheet["C18"] = str(total_cbal)
    sheet["C19"] = str(cbal_count)
    sheet["D18"] = str(diff_cbal)
    sheet["D19"] = str(diff_count)
    sheet["E18"] = str(different_cbal)
    sheet["E19"] = str(different_count)
    sheet["C17"] = "Loan Level"
    sheet["D17"] = "Difference"
    sheet["E17"] = "Difference %"
    sheet["A18"] = "CBAL"
    sheet["A19"] = "COUNT*"
    regular = Side(border_style="medium", color="000000")
    no_left_side = Border(top = regular,bottom=regular,right=regular)
    for c in sheet['A17:A19']+sheet['B17:B19']+sheet['C17:C19']+sheet['D17:D19']+sheet['E17:E19']:
        c[0].border = no_left_side
    for b in sheet['A17:A19']+sheet['B15:B17']+sheet['C17:C17']+sheet['D17:D17']+sheet['E17:E17']:
        b[0].font = Font(bold=True)
else:
    pass
if Deal_type=="First Performance":
    sheet["A15"] = "Period – Sep 2021"
    sheet["B15"] = "First Performance"
    sheet["A17"] = ""
    sheet["B17"] = "XML"
    sheet["B18"] = float(CBAL)
    sheet["B19"] = int(COUNT)
    sheet["C18"] = str(total_cbal)
    sheet["C19"] = str(cbal_count)
    sheet["D18"] = str(diff_cbal)
    sheet["D19"] = str(diff_count)
    sheet["E18"] = str(different_cbal)
    sheet["E19"] = str(different_count)
    sheet["C17"] = "Loan Level"
    sheet["D17"] = "Difference"
    sheet["E17"] = "Difference %"
    sheet["A18"] = "CBAL"
    sheet["A19"] = "COUNT*"
    regular = Side(border_style="medium", color="000000")
    no_left_side = Border(top = regular,bottom=regular,right=regular)
    for c in sheet['A17:A19']+sheet['B17:B19']+sheet['C17:C19']+sheet['D17:D19']+sheet['E17:E19']:
        c[0].border = no_left_side
    for b in sheet['A17:A19']+sheet['B15:B17']+sheet['C17:C17']+sheet['D17:D17']+sheet['E17:E17']:
        b[0].font = Font(bold=True)
else:
    pass
sheet["A21"] = "Note: "
sheet["A23"] = "As we can see that there no difference between Loan Level and Aggregate CBAL,\n we can consider the mapping as cutoffdt equals asoffdt, let me know your thoughts."
sheet["A24"] = "For CBAL and Count exc RED loans"
#Fill the cell with color
color_fill = PatternFill(start_color='FFFA2A', fill_type='solid')
sheet['A1'].fill = color_fill #assign the column want to fill with color
sheet["A1"].alignment = Alignment(horizontal='center') #cell to be centered

double = Side(border_style="double", color="000000")
thin = Side(border_style="medium", color="000000")
regular = Side(border_style="medium", color="000000")

## For the title cells B2 to F2
for c in sheet['A1:B1'][0]:
    c.border = Border(bottom=double, top=double)

no_left_side = Border(top = regular,bottom=regular,right=regular)
no_right_side = Border(top = regular,bottom=regular, left=regular)
box = Border(top = regular,bottom=regular, left=regular,right=regular)

## For the "table-like" cells
for c in sheet['A2:A13']+sheet['B2:B13']:
    c[0].border = no_left_side
    
sheet['A35'].alignment = Alignment(wrapText=True)
sheet['A35'].alignment = Alignment(horizontal='left', vertical ='center')
sheet['A1'].font = Font(bold=True)

for column_cells in sheet.columns:
    new_column_length = max(len(str(cell.value)) for cell in column_cells)
    new_column_letter = (get_column_letter(column_cells[0].column))
    if new_column_length > 0:
        sheet.column_dimensions[new_column_letter].width = new_column_length*0.5
        
        
for b in sheet['A2:A21']:
    b[0].font = Font(bold=True)
    

file = your_workbook.save("C:\\Users\\abhijeet_gode\\Downloads\\Bond mapping.xlsx") #saving the file with the 'xlsx' excel extension
print('Program is successfully executed.................................. :)')